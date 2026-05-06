#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IZRACUNAVANJE KPIs IZ FORCE PLATE PODATAKA
==========================================
Koristi logiku iz VJ_1_1.py za izračunavanje KPIs iz Force Plate fajlova
i dodaje nove listove SJ_FP i CMJ_FP u postojeći Excel fajl.
"""

import sys
import re
import numpy as np
import pandas as pd
from pathlib import Path
from scipy import signal
from scipy.integrate import cumulative_trapezoid
from datetime import datetime

# Konfiguracija (iz VJ_1_1.py)
CONFIG = {
    'GRAVITY': 9.81,
    'FILTER_FREQ': 50,       # Hz
    'FILTER_ORDER': 2,       # Butterworth Order
    'ONSET_THRESHOLD_SD': 5, # Standard Deviations for onset
    'FLIGHT_THRESHOLD': 15,  # Newtons
    'CORRECT_DRIFT': True,
    'DRIFT_ZERO_PERIOD': 0.5, # Seconds at end of file for slope calc
    'SAMPLE_RATE_AMTI': 1000.0,
}


def _mad(x):
    """Median Absolute Deviation."""
    x = np.asarray(x, float)
    med = np.nanmedian(x)
    return np.nanmedian(np.abs(x - med))


def _robust_sigma_mad(x):
    """Robust SD estimate: 1.4826 * MAD."""
    return 1.4826 * _mad(x)


def _first_sustained(mask, min_len):
    """Return first index i such that mask[i:i+min_len] are all True."""
    mask = np.asarray(mask, bool)
    if mask.size < min_len:
        return None
    run = np.convolve(mask.astype(int), np.ones(min_len, dtype=int), mode="valid")
    hit = np.where(run == min_len)[0]
    return int(hit[0]) if hit.size else None


def _pick_quiet_baseline(force, fs, start_s=0.2, win_s=1.0, search_end_s=2.2):
    """Pick the quietest baseline window by minimizing RMS around the median."""
    f = np.asarray(force, float)
    n = f.size
    i0 = int(round(start_s * fs))
    i_end = int(round(min(search_end_s * fs, n - 1)))
    w = int(round(win_s * fs))
    w = max(w, 50)

    if i_end - i0 < w + 5:
        return max(0, i0), min(n, i0 + w)

    best_i = i0
    best_val = np.inf
    for i in range(i0, i_end - w):
        seg = f[i:i + w]
        if not np.isfinite(seg).all():
            continue
        m = np.median(seg)
        rms = np.sqrt(np.mean((seg - m) ** 2))
        if rms < best_val:
            best_val = rms
            best_i = i

    return best_i, best_i + w


def detect_fp_onset_unweighting(
    force,
    fs=1000,
    baseline_start_s=0.2,
    baseline_win_s=1.0,
    baseline_search_end_s=2.2,
    k_sigma=5.0,
    floor_frac_bw=0.02,
    floor_N=20.0,
    persist_s=0.04,
):
    """
    Robust force-plate CMJ onset (start of unweighting):
    earliest time after baseline where Fz < BW - T for >= persist_s.
    T = max(k_sigma*sigma, floor_frac_bw*BW, floor_N)
    """
    f = np.asarray(force, float)
    n = f.size
    if n < int(fs * 0.5):
        return None

    b0, b1 = _pick_quiet_baseline(
        f, fs,
        start_s=baseline_start_s,
        win_s=baseline_win_s,
        search_end_s=baseline_search_end_s
    )
    base = f[b0:b1]
    BW = float(np.median(base))

    fnet_base = base - BW
    sigma = float(_robust_sigma_mad(fnet_base))

    T = max(k_sigma * sigma, floor_frac_bw * BW, floor_N)

    persist = max(3, int(round(persist_s * fs)))
    start_idx = b1

    mask = np.zeros(n, dtype=bool)
    mask[start_idx:] = f[start_idx:] < (BW - T)

    idx = _first_sustained(mask, persist)
    if idx is None:
        return None

    return idx


def butter_lowpass_filter(data, cutoff, fs, order=2):
    """Apply a zero-lag Butterworth low-pass filter."""
    nyquist = 0.5 * fs
    normal_cutoff = cutoff / nyquist
    b, a = signal.butter(order, normal_cutoff, btype='low', analog=False)
    if len(data) <= 3 * max(len(b), len(a)): 
        return data 
    return signal.filtfilt(b, a, data)


def calculate_body_weight_robust(force, fs, min_idx):
    """
    Robustniji algoritam za izračunavanje body weight-a.
    POBOLJŠANJE v2: Prvo nađe period sa kontaktom, pa računa BW iz tog perioda.
    """
    n_samples = len(force)
    
    # Validacija
    if n_samples < 100:
        return calculate_body_weight(force, fs)
    
    # STRATEGIJA: Prvo nađi period gde je kontakt sa Force Plate-om
    # Kontakt = sila > threshold (npr. 100N)
    CONTACT_THRESHOLD = 100.0  # Minimum sila za kontakt
    
    # Traži period sa kontaktom u prvih 3s signala
    search_end = min(int(3.0 * fs), n_samples)
    contact_mask = force[:search_end] > CONTACT_THRESHOLD
    
    if np.sum(contact_mask) < 100:
        # Nema dovoljno kontakta - možda je signal loš
        # Probaj sa nižim threshold-om
        CONTACT_THRESHOLD = 50.0
        contact_mask = force[:search_end] > CONTACT_THRESHOLD
        
        if np.sum(contact_mask) < 100:
            # Još uvek nema kontakta - koristi fallback
            return calculate_body_weight(force, fs)
    
    # Nađi prvi i poslednji sample sa kontaktom u prvih 3s
    contact_indices = np.where(contact_mask)[0]
    if len(contact_indices) == 0:
        return calculate_body_weight(force, fs)
    
    # Dodatna provera da li ima dovoljno kontakta
    if len(contact_indices) < 100:
        return calculate_body_weight(force, fs)
    
    first_contact = contact_indices[0]
    last_contact = contact_indices[-1]
    
    # Koristi period sa kontaktom za BW (quiet standing)
    # Uzmi period od prvog kontakta + 0.1s do prvog kontakta + 1.5s (ili do poslednjeg kontakta)
    quiet_period_start = first_contact + int(0.1 * fs)  # Preskoči prvih 100ms nakon kontakta
    quiet_period_end = min(first_contact + int(1.5 * fs), last_contact + int(0.1 * fs), n_samples)
    
    # Proveri da li je period validan
    if quiet_period_end <= quiet_period_start or quiet_period_end - quiet_period_start < 100:
        # Period je previše kratak - koristi period od prvog kontakta
        quiet_period_start = first_contact
        quiet_period_end = min(first_contact + int(1.0 * fs), n_samples)
        
        if quiet_period_end <= quiet_period_start or quiet_period_end - quiet_period_start < 100:
            return calculate_body_weight(force, fs)
    
    quiet_segment = force[quiet_period_start:quiet_period_end]
    
    # Proveri da li je segment razuman (sila treba biti > 100N)
    if np.mean(quiet_segment) < CONTACT_THRESHOLD:
        # Segment nema dovoljno kontakta - koristi fallback
        return calculate_body_weight(force, fs)
    
    try:
        # Prvo aproksimacija
        bw_est = np.median(quiet_segment)
        bw_std_est = np.std(quiet_segment, ddof=1)
        
        if np.isnan(bw_est) or np.isnan(bw_std_est) or bw_std_est == 0:
            return calculate_body_weight(force, fs)
        
        # Filtriranje: tačke unutar 2*SD (ukloni artefakte i ekstremne vrednosti)
        mask = (quiet_segment > bw_est - 2 * bw_std_est) & (quiet_segment < bw_est + 2 * bw_std_est)
        clean_indices = np.where(mask)[0]
        
        if len(clean_indices) > 100:
            final_bw = np.mean(quiet_segment[clean_indices])
            final_sd = np.std(quiet_segment[clean_indices], ddof=1)
        else:
            # Ako nema dovoljno čistih tačaka, koristi ceo segment
            final_bw = np.mean(quiet_segment)
            final_sd = np.std(quiet_segment, ddof=1)
        
        if np.isnan(final_bw) or np.isnan(final_sd) or final_sd == 0:
            return calculate_body_weight(force, fs)
        
        # VALIDACIJA: Proveri da li je BW razuman (300-1500N za odraslu osobu)
        # Strožija validacija
        if final_bw < 300 or final_bw > 1500:
            # Fallback na staru metodu
            return calculate_body_weight(force, fs)
        
        return final_bw, final_sd
    except:
        # Fallback na staru metodu
        return calculate_body_weight(force, fs)


def calculate_body_weight(force, fs):
    """Robust iterative algorithm to find quiet standing weight (fallback metoda)."""
    n_samples = len(force)
    end_init = min(3000, n_samples)
    slice_init = force[99:end_init] if n_samples > 100 else force
    bw_est = np.mean(slice_init)
    bw_std = np.std(slice_init, ddof=1)
    
    search_end = min(3500, n_samples)
    search_slice = force[99:search_end] if n_samples > 100 else force
    mask = (search_slice > bw_est - bw_std) & (search_slice < bw_est + bw_std)
    clean_indices = np.where(mask)[0]
    
    if len(clean_indices) > 50:
        final_bw = np.mean(search_slice[clean_indices])
        final_sd = np.std(search_slice[clean_indices], ddof=1)
    else:
        final_bw = bw_est
        final_sd = bw_std
        
    return final_bw, final_sd


def read_force_file(filepath):
    """Read AMTI force plate file."""
    try:
        # Pokušaj sa različitim separatorima/encoding opcijama
        # Neki AMTI fajlovi imaju latin-1 znakove ili nemaju header red.
        read_attempts = [
            {'skiprows': 1, 'header': None, 'sep': None, 'engine': 'python', 'decimal': ',', 'encoding': 'utf-8'},
            {'skiprows': 1, 'header': None, 'sep': ',', 'decimal': ',', 'encoding': 'utf-8'},
            {'skiprows': 0, 'header': None, 'sep': None, 'engine': 'python', 'decimal': ',', 'encoding': 'utf-8'},
            {'skiprows': 0, 'header': None, 'sep': ',', 'decimal': ',', 'encoding': 'utf-8'},
            {'skiprows': 1, 'header': None, 'sep': None, 'engine': 'python', 'decimal': ',', 'encoding': 'latin1'},
            {'skiprows': 1, 'header': None, 'sep': ',', 'decimal': ',', 'encoding': 'latin1'},
            {'skiprows': 0, 'header': None, 'sep': None, 'engine': 'python', 'decimal': ',', 'encoding': 'latin1'},
            {'skiprows': 0, 'header': None, 'sep': ',', 'decimal': ',', 'encoding': 'latin1'},
        ]

        df = None
        last_error = None
        for kwargs in read_attempts:
            try:
                df_candidate = pd.read_csv(filepath, **kwargs)
                if df_candidate is not None and not df_candidate.empty and df_candidate.shape[1] >= 2:
                    df = df_candidate
                    break
            except Exception as e:
                last_error = e
                continue
        
        if df is None:
            if last_error is not None:
                raise last_error
            raise ValueError("No valid data columns found")
        
        if df.shape[1] >= 9:
            col_L, col_R = 2, 8
        else:
            col_L, col_R = 0, 1
        
        raw_L = pd.to_numeric(df.iloc[:, col_L], errors='coerce').values
        raw_R = pd.to_numeric(df.iloc[:, col_R], errors='coerce').values
        
        valid = ~np.isnan(raw_L) & ~np.isnan(raw_R)
        return raw_L[valid], raw_R[valid]
    
    except Exception as e:
        print(f"  [Read Error] {Path(filepath).name}: {e}")
        return None, None


def detect_onset_robust(force, bw, bw_sd, fs, search_end_idx, jump_type):
    """
    Robust onset detection with sustained threshold crossing.
    - CMJ: onset when force starts sustained unloading (below BW band)
    - SJ: onset when force starts sustained loading (above BW band)
    """
    n = len(force)
    if n == 0:
        return 0
    search_end_idx = int(np.clip(search_end_idx, 1, n - 1))

    min_stable_period = int(0.30 * fs)
    start_idx = min(min_stable_period, max(0, search_end_idx - 1))
    if search_end_idx <= start_idx + 5:
        start_idx = 0

    threshold_n = max(2.5 * bw_sd, 15.0)
    baseline_band = max(1.5 * bw_sd, 10.0)
    sustain_n = max(int(0.020 * fs), 8)  # ~20 ms
    back_window = max(int(0.250 * fs), sustain_n)

    if jump_type == 2:  # CMJ -> unloading first
        trigger_mask = force[start_idx:search_end_idx] < (bw - threshold_n)
    else:  # SJ -> loading first
        trigger_mask = force[start_idx:search_end_idx] > (bw + threshold_n)

    if len(trigger_mask) < sustain_n:
        return start_idx

    # Find first sustained trigger period
    trigger_int = trigger_mask.astype(int)
    rolling = np.convolve(trigger_int, np.ones(sustain_n, dtype=int), mode='valid')
    sustained_hits = np.where(rolling >= sustain_n)[0]
    if len(sustained_hits) == 0:
        return start_idx

    crossing_abs = start_idx + int(sustained_hits[0])

    # Backtrack to last baseline sample right before movement starts
    b0 = max(start_idx, crossing_abs - back_window)
    baseline_slice = force[b0:crossing_abs + 1]
    in_baseline = (baseline_slice >= (bw - baseline_band)) & (baseline_slice <= (bw + baseline_band))
    baseline_idxs = np.where(in_baseline)[0]
    if len(baseline_idxs) > 0:
        onset_abs = b0 + int(baseline_idxs[-1])
    else:
        onset_abs = crossing_abs

    return int(np.clip(onset_abs, 0, n - 1))


def analyze_jump(f_tot_raw, f_L_raw, f_R_raw, fs, jump_type, file_id, return_timeseries=False):
    """Perform full biomechanical analysis on a single trial (from VJ_1_1.py)."""
    
    # QC Flags - inicijalizuj na početku
    qc_flags = {
        'has_countermovement': False,
        'negative_vto': False,
        'invalid_events': False,
        'invalid_bw': False,
        'notes': ''
    }
    
    # A. Filtering
    force = butter_lowpass_filter(f_tot_raw, CONFIG['FILTER_FREQ'], fs, CONFIG['FILTER_ORDER'])
    force_L = butter_lowpass_filter(f_L_raw, CONFIG['FILTER_FREQ'], fs, CONFIG['FILTER_ORDER'])
    force_R = butter_lowpass_filter(f_R_raw, CONFIG['FILTER_FREQ'], fs, CONFIG['FILTER_ORDER'])
    
    # B. Nađi apsolutni minimum u celom signalu (flight faza) - PRE CROPOVANJA
    idx_abs_min = np.argmin(force)
    
    # C. Robustniji Body Weight iz stabilnog perioda pre minimuma
    bw, bw_sd = calculate_body_weight_robust(force, fs, idx_abs_min)
    
    # Proveri da li je BW razuman - dodatna validacija
    if bw < 300 or bw > 1500 or np.isnan(bw) or np.isnan(bw_sd) or bw_sd <= 0:
        qc_flags['invalid_bw'] = True
        if qc_flags['notes']:
            qc_flags['notes'] += f'; Invalid BW ({bw:.2f}N)'
        else:
            qc_flags['notes'] = f'Invalid BW ({bw:.2f}N)'
        # Ako je BW potpuno nerealan, ne možemo nastaviti
        if bw < 100 or bw > 2000:
            return None
    bm = bw / CONFIG['GRAVITY']
    
    # D. Nađi pike oko apsolutnog minimuma u ORIGINALNOM signalu
    # Pronađi maksimum LEVO od minimuma (propulsion peak)
    if idx_abs_min > 0:
        left_segment = force[:idx_abs_min]
        if len(left_segment) > 0:
            propulsion_peak_abs = np.argmax(left_segment)
        else:
            propulsion_peak_abs = 0
    else:
        propulsion_peak_abs = 0
    
    # Pronađi maksimum DESNO od minimuma (landing peak)
    if idx_abs_min < len(force) - 1:
        right_segment = force[idx_abs_min + 1:]
        if len(right_segment) > 0:
            landing_peak_abs = idx_abs_min + 1 + np.argmax(right_segment)
        else:
            landing_peak_abs = len(force) - 1
    else:
        landing_peak_abs = len(force) - 1
    
    # Validacija: proveri da li su pike validni
    if propulsion_peak_abs >= len(force) or landing_peak_abs >= len(force) or propulsion_peak_abs < 0 or landing_peak_abs < 0:
        qc_flags['invalid_events'] = True
        qc_flags['notes'] = 'Invalid peak detection'
        return None
    
    # Validacija: proveri da li je minimum između pika
    if not (propulsion_peak_abs < idx_abs_min < landing_peak_abs):
        qc_flags['invalid_events'] = True
        qc_flags['notes'] = 'Invalid peak-minimum relationship'
        return None
    
    # Precizna detekcija take-off (F) u ORIGINALNOM signalu
    # Od propulsion peak-a, idi DESNO i nađi prvu tačku gde je sila < 50N
    idx_F_abs = propulsion_peak_abs
    for i in range(propulsion_peak_abs, min(len(force), propulsion_peak_abs + int(0.5 * fs))):
        if force[i] < 50.0:
            idx_F_abs = i
            # Proveri: ako je blizu (unutar 10 samples) veća od 10N, pomeri za još jednu tačku desno
            if i + 1 < len(force) and force[i + 1] > 10.0:
                idx_F_abs = i + 1
            break
    
    # Precizna detekcija landing (H) u ORIGINALNOM signalu
    # Od landing peak-a, idi LEVO i nađi prvu tačku gde je sila < 50N
    idx_H_abs = landing_peak_abs
    for i in range(landing_peak_abs, max(0, landing_peak_abs - int(0.5 * fs)), -1):
        if force[i] < 50.0:
            idx_H_abs = i
            # Proveri: ako je blizu veća od 10N, pomeri za još jednu tačku levo
            if i - 1 >= 0 and force[i - 1] > 10.0:
                idx_H_abs = i - 1
            break
    
    # Onset (A) - robust sustained-threshold logika
    idx_A_abs = 0
    
    if jump_type == 1:  # SJ (Squat Jump)
        # Nađi maksimum od početka signala do minimuma
        if idx_abs_min > 0 and idx_abs_min < len(force):
            segment_to_min = force[:idx_abs_min]
            if len(segment_to_min) > 0:
                max_before_min_abs = np.argmax(segment_to_min)
            else:
                max_before_min_abs = 0
        else:
            max_before_min_abs = 0
        
        # Od maksimuma idi unazad (levo) i traži prvu tačku gde je sila < BW + 5*SD
        threshold_sj = bw + (5 * bw_sd)
        idx_A_abs = max_before_min_abs
        
        for i in range(max_before_min_abs, -1, -1):
            if force[i] < threshold_sj:
                idx_A_abs = i
                break
        
        # NOVA LOGIKA: Detekcija countermovement-a koristeći CMJ pristup
        # Ako postoji countermovement, postoji unweighting minimum (kao u CMJ)
        # Koristi unweighting minimum za detektovanje pravog početka - backward od te tačke
        
        cm_threshold = bw - (5 * bw_sd)  # Threshold za detekciju countermovement-a (5*SD jer je SD mali ako je BW dobro određen)
        
        # Proveri da li postoji countermovement (sila < BW - 2*SD) pre maksimuma
        # VAŽNO: Ignoriši countermovement na samom početku signala (pre stabilizacije)
        min_stable_period = int(0.5 * fs)  # Minimum period stabilnosti pre traženja CM
        
        if max_before_min_abs > min_stable_period:
            # Traži countermovement samo posle perioda stabilnosti
            segment_before_max = force[min_stable_period:max_before_min_abs]
            cm_indices = np.where(segment_before_max < cm_threshold)[0]
            
            if len(cm_indices) > 0:
                # Postoji countermovement!
                qc_flags['has_countermovement'] = True
                
                # POBOLJŠANJE: Koristi CMJ logiku - nađi unweighting minimum
                # Unweighting minimum je minimum između početka i maksimuma (max_before_min_abs)
                if max_before_min_abs > min_stable_period:
                    segment_to_max = force[min_stable_period:max_before_min_abs]
                    if len(segment_to_max) > 0:
                        unweighting_min_rel = np.argmin(segment_to_max)
                        unweighting_min_abs = unweighting_min_rel + min_stable_period
                    else:
                        unweighting_min_abs = min_stable_period
                else:
                    unweighting_min_abs = min_stable_period
                
                # Od unweighting minimuma idi unazad (ka početku) i nađi pravi početak
                # POBOLJŠANJE: Traži stabilnu tačku pre početka pada (slično kao u CMJ, ali sa dodatnom proverom)
                threshold_cmj = bw - (5 * bw_sd)
                stable_threshold_low = bw - (0.5 * bw_sd)
                stable_threshold_high = bw + (0.5 * bw_sd)
                
                # Traži unazad od unweighting/unloading minimuma do početka signala
                # Prvo nađi prvu tačku gde je F >= BW - 5*SD
                idx_A_abs = unweighting_min_abs
                found_threshold = False
                for i in range(unweighting_min_abs, -1, -1):
                    if force[i] >= threshold_cmj:
                        idx_A_abs = i
                        found_threshold = True
                        break
                
                # POBOLJŠANJE: Ako smo našli threshold tačku, proveri da li je stabilna
                # Traži unazad od threshold tačke i nađi poslednju stabilnu tačku (BW ± 0.5*SD)
                # pre nego što sila počinje da pada
                if found_threshold and idx_A_abs > min_stable_period:
                    # Traži unazad od threshold tačke do min_stable_period
                    search_start = max(min_stable_period, idx_A_abs - int(1.0 * fs))
                    true_start_idx = None
                    
                    for i in range(idx_A_abs - 1, search_start - 1, -1):
                        if stable_threshold_low <= force[i] <= stable_threshold_high:
                            # Našli smo stabilnu tačku - proveri da li je ovo početak pada
                            if i + 1 < len(force) and force[i + 1] < force[i]:
                                true_start_idx = i
                                # Proveri da li je ovo dovoljno stabilna tačka
                                check_window = min(10, i)
                                if check_window > 0:
                                    prev_window = force[i - check_window:i]
                                    prev_stable = np.sum((stable_threshold_low <= prev_window) & (prev_window <= stable_threshold_high))
                                    if prev_stable >= check_window * 0.7:  # Bar 70% prethodnih tačaka je stabilno
                                        break
                                else:
                                    break
                    
                    # Ako nismo našli tačku gde sila počinje da pada, traži poslednju stabilnu tačku
                    if true_start_idx is None:
                        for i in range(idx_A_abs - 1, search_start - 1, -1):
                            if stable_threshold_low <= force[i] <= stable_threshold_high:
                                true_start_idx = i
                                break
                    
                    if true_start_idx is not None:
                        idx_A_abs = true_start_idx
                
                # Ako nismo našli (što ne bi trebalo da se desi), koristi početak signala
                if idx_A_abs == unweighting_min_abs:
                    # Fallback: koristi početak signala ili min_stable_period
                    idx_A_abs = max(0, min_stable_period)
                
                qc_flags['notes'] = f'Countermovement detected: unweighting min @ sample {unweighting_min_abs}, onset @ sample {idx_A_abs}'
            else:
                # Nema countermovement-a - koristi standardnu SJ logiku
                pass
        
    else:  # CMJ (Countermovement Jump)
        # ROBUST ONSET: MAD-based, persistence, threshold floors
        robust_onset = detect_fp_onset_unweighting(force, fs=fs)
        if robust_onset is not None:
            idx_A_abs = robust_onset
        else:
            # Fallback na staru logiku ako robust ne nadje onset
            if idx_abs_min > 0 and idx_abs_min < len(force):
                segment_to_min = force[:idx_abs_min]
                propulsion_peak_abs = np.argmax(segment_to_min) if len(segment_to_min) > 0 else 0
            else:
                propulsion_peak_abs = 0
            if propulsion_peak_abs > 0:
                segment_to_peak = force[:propulsion_peak_abs]
                unweighting_min_abs = np.argmin(segment_to_peak) if len(segment_to_peak) > 0 else 0
            else:
                unweighting_min_abs = 0
            threshold_cmj = bw - (5 * bw_sd)
            idx_A_abs = unweighting_min_abs
            for i in range(unweighting_min_abs, -1, -1):
                if force[i] >= threshold_cmj:
                    idx_A_abs = i
                    break
            if idx_A_abs == unweighting_min_abs:
                idx_A_abs = 0

        # Unweighting min (B): između onset-a i propulsion peaka
        if idx_abs_min > 0 and idx_abs_min < len(force):
            segment_to_min = force[:idx_abs_min]
            propulsion_peak_abs = np.argmax(segment_to_min) if len(segment_to_min) > 0 else 0
        else:
            propulsion_peak_abs = 0
        if propulsion_peak_abs > idx_A_abs:
            segment_between = force[idx_A_abs:propulsion_peak_abs]
            unweighting_min_abs = idx_A_abs + np.argmin(segment_between) if len(segment_between) > 0 else idx_A_abs
        else:
            unweighting_min_abs = idx_A_abs

    # Zavrsna validacija za onset
    # Mora biti pre takeoff-a i pre unweighting minimuma (ako postoji)
    if idx_A_abs >= idx_F_abs:
        idx_A_abs = max(0, idx_F_abs - int(0.05 * fs))
    if 'unweighting_min_abs' in locals() and unweighting_min_abs is not None and idx_A_abs >= unweighting_min_abs:
        idx_A_abs = max(0, unweighting_min_abs - int(0.03 * fs))
    
    # E. Cropping oko skoka (za dalju analizu i integraciju)
    # Cropuj oko onset-a i landing-a sa padding-om
    pad_pre, pad_post = 1000, 1500
    start_crop = max(0, idx_A_abs - pad_pre)
    end_crop = min(len(force), idx_H_abs + pad_post)
    
    f_crop = force[start_crop:end_crop]
    fL_crop = force_L[start_crop:end_crop]
    fR_crop = force_R[start_crop:end_crop]
    t_crop = np.arange(len(f_crop)) / fs
    
    # Prilagodi indekse za crop-ovani signal
    idx = {}
    idx['A'] = idx_A_abs - start_crop
    idx['F'] = idx_F_abs - start_crop
    idx['H'] = idx_H_abs - start_crop
    idx['min'] = idx_abs_min - start_crop
    idx['propulsion_peak'] = propulsion_peak_abs - start_crop
    idx['landing_peak'] = landing_peak_abs - start_crop
    
    # Za CMJ: Dodaj unweighting_min_abs u idx mapiranje
    # Za SJ sa CM: Takođe dodaj unweighting_min_abs ako postoji
    if jump_type == 2:  # CMJ
        # unweighting_min_abs je već izračunat u linijama 408-416
        # Proveri da li je između onset-a i propulsion peak-a
        if idx_A_abs <= unweighting_min_abs < propulsion_peak_abs:
            idx['B'] = unweighting_min_abs - start_crop
        else:
            # Rekalkuliši između onset-a i propulsion peak-a
            if idx_A_abs < propulsion_peak_abs:
                segment_between = force[idx_A_abs:propulsion_peak_abs]
                if len(segment_between) > 0:
                    unweighting_min_rel = np.argmin(segment_between)
                    unweighting_min_abs_recalc = idx_A_abs + unweighting_min_rel
                    idx['B'] = unweighting_min_abs_recalc - start_crop
                else:
                    idx['B'] = -1
            else:
                idx['B'] = -1
    else:  # SJ
        # Za SJ, proveri da li postoji unweighting_min_abs iz CM detekcije
        # (izračunat u linijama 324-329)
        if 'unweighting_min_abs' in locals() and unweighting_min_abs is not None:
            if idx_A_abs <= unweighting_min_abs < propulsion_peak_abs:
                idx['B'] = unweighting_min_abs - start_crop
            else:
                idx['B'] = -1
        else:
            idx['B'] = -1
    
    # E. Integration (Kinematics)
    acc = (f_crop - bw) / bm
    
    # Velocity
    # VAŽNO: Forward integracija od onset-a sa initial=0 garantuje da brzina počinje od nule
    # Proveri da li postoji countermovement za SJ
    has_cm = False
    if jump_type == 1:
        # Proveri da li je detektovan countermovement (iz qc_flags ili iz idx['B'])
        if qc_flags.get('has_countermovement', False) or idx.get('B', -1) >= 0:
            has_cm = True
    
    if jump_type == 1 and has_cm and idx['A'] >= 0:
        # Za SJ SA COUNTERMOVEMENT-OM: integracija počinje direktno od onset-a
        acc_segment = acc[idx['A']:]
        t_segment = t_crop[idx['A']:] - t_crop[idx['A']]  # Relativno vreme od onset-a
        if len(acc_segment) > 0 and len(t_segment) > 0:
            vel_segment = cumulative_trapezoid(acc_segment, t_segment, initial=0)
            vel = np.zeros_like(acc)
            vel[idx['A']:] = vel_segment
            # EKSPLICITNO: Osiguraj da je brzina na onset-u tačno 0
            if idx['A'] >= 0 and idx['A'] < len(vel):
                vel[idx['A']] = 0.0
        else:
            # Fallback: standard forward integration
            vel = cumulative_trapezoid(acc, t_crop, initial=0)
            if idx['A'] >= 0 and idx['A'] < len(vel):
                vel = vel - vel[idx['A']]
                vel[idx['A']] = 0.0
    else:
        # Za CMJ ili SJ BEZ countermovement-a: standardna integracija (ISTA LOGIKA KAO U compare_velocity_profiles.py)
        vel = cumulative_trapezoid(acc, t_crop, initial=0)
        if idx['A'] >= 0 and idx['A'] < len(vel):
            vel = vel - vel[idx['A']]  # Oduzmi offset da bi brzina na onset-u bila 0
            vel[idx['A']] = 0.0  # Eksplicitno postavi na 0
        
    # Drift Correction - POBOLJŠANJE
    if CONFIG.get('CORRECT_DRIFT', True):
        # Koristi landing period (posle impact) za drift correction umesto kraja signala
        # Ovo je pouzdanije jer je landing period stabilniji
        landing_buffer = int(CONFIG.get('DRIFT_ZERO_PERIOD', 0.5) * fs)
        
        # Proveri da li imamo dovoljno podataka posle landing-a
        if idx['H'] + landing_buffer < len(vel):
            # Koristi period posle landing-a (impact) za drift estimation
            v_landing_observed = np.mean(vel[idx['H']:idx['H'] + landing_buffer])
            duration = t_crop[idx['H']] - t_crop[idx['A']] if idx['H'] > idx['A'] else t_crop[-1] - t_crop[idx['A']]
            
            if duration > 0 and abs(v_landing_observed) < 10.0:  # Validacija - brzina na landing-u treba biti blizu 0
                drift_slope = v_landing_observed / duration
                correction = np.zeros_like(vel)
                correction[idx['A']:] = drift_slope * (t_crop[idx['A']:] - t_crop[idx['A']])
                vel = vel - correction
                # EKSPLICITNO: Osiguraj da je brzina na onset-u tačno 0 i posle drift correction
                if idx['A'] >= 0 and idx['A'] < len(vel):
                    vel[idx['A']] = 0.0
            else:
                # Fallback: koristi kraj signala samo ako landing period nije validan
                end_buffer = int(0.3 * fs)  # Kraći buffer za fallback
                if len(vel) > idx['H'] + end_buffer:
                    v_end_observed = np.mean(vel[-end_buffer:])
                    duration = t_crop[-1] - t_crop[idx['A']]
                    if duration > 0 and abs(v_end_observed) < 5.0:  # Validacija
                        drift_slope = v_end_observed / duration
                        correction = np.zeros_like(vel)
                        correction[idx['A']:] = drift_slope * (t_crop[idx['A']:] - t_crop[idx['A']])
                        vel = vel - correction
                        # EKSPLICITNO: Osiguraj da je brzina na onset-u tačno 0 i posle drift correction
                        if idx['A'] >= 0 and idx['A'] < len(vel):
                            vel[idx['A']] = 0.0
    
    # Displacement
    disp = cumulative_trapezoid(vel, t_crop, initial=0)
    if idx['A'] < len(disp):
        disp = disp - disp[idx['A']]
    
    power = f_crop * vel
    
    # F. Phase Landmarks
    slice_end = idx['F']
    if idx['A'] < slice_end and slice_end <= len(vel):
        vel_segment = vel[idx['A']:slice_end]
        if len(vel_segment) > 0:
            idx['E'] = idx['A'] + np.argmax(vel_segment)
        else:
            idx['E'] = idx['A']
    else:
        idx['E'] = idx['A']
    
    # QC Flags - inicijalizuj pre detekcije događaja
    qc_flags = {
        'has_countermovement': False,
        'negative_vto': False,
        'invalid_events': False,
        'notes': ''
    }
    
    if jump_type != 1:  # CMJ
        # VAŽNO: Redosled događaja za CMJ:
        # A = Onset, B = Unweighting min, C = Propulsion peak, D = Takeoff, E = Peak velocity, F = Takeoff (isti kao D)
        # idx['C'] treba da bude propulsion peak (maksimum sile), ne minimum brzine!
        idx['C'] = idx.get('propulsion_peak', idx['A'])  # Propulsion peak (maksimum sile pre TO)
        
        # idx['D'] je takeoff - koristi idx['F'] koji je već izračunat
        idx['D'] = idx['F']  # Takeoff 
    else:  # SJ - DETECT COUNTERMOVEMENT
        idx['C'] = idx['A']
        idx['D'] = idx['A']
        
        # Detekcija countermovement za SJ:
        # 1. Proveri da li postoji značajno spuštanje (negativna brzina) pre takeoff
        # 2. Proveri da li postoji značajno spuštanje displacement-a
        if idx['A'] < idx['F']:
            vel_pre_to = vel[idx['A']:idx['F']]
            disp_pre_to = disp[idx['A']:idx['F']]
            
            # Proveri negativnu brzinu (spuštanje)
            min_vel = np.min(vel_pre_to) if len(vel_pre_to) > 0 else 0
            min_disp = np.min(disp_pre_to) if len(disp_pre_to) > 0 else 0
            
            # Kriterijumi za countermovement (POOŠTRAVANJE):
            # - Minimalna brzina < -0.15 m/s (značajno spuštanje, ne samo šum)
            # - Minimalni displacement < -0.02 m (20mm spuštanje, ne samo šum)
            # - Minimalna sila < BW - 3*SD (značajno smanjenje sile, ne samo varijacija)
            # - Trajanje countermovement-a > 50ms (ne samo kratkotrajni šum)
            f_pre_to = f_crop[idx['A']:idx['F']]
            min_force = np.min(f_pre_to) if len(f_pre_to) > 0 else bw
            
            cm_vel_threshold = -0.15  # m/s (pooštravanje sa -0.1)
            cm_disp_threshold = -0.02  # m (20mm, pooštravanje sa 10mm)
            cm_force_threshold = bw - (3 * bw_sd)  # N (pooštravanje sa 2*SD)
            
            # Proveri trajanje countermovement-a
            cm_duration_threshold = 0.05  # s (50ms)
            cm_duration = 0.0
            
            if min_vel < cm_vel_threshold:
                # Pronađi kontinuirani period negativne brzine
                vel_below_threshold = vel_pre_to < cm_vel_threshold
                if np.any(vel_below_threshold):
                    # Pronađi najduži kontinuirani period
                    diff_vel = np.diff(np.concatenate(([False], vel_below_threshold, [False])).astype(int))
                    starts = np.where(diff_vel == 1)[0]
                    ends = np.where(diff_vel == -1)[0]
                    if len(starts) > 0 and len(ends) > 0:
                        durations = [(ends[i] - starts[i]) / fs for i in range(min(len(starts), len(ends)))]
                        cm_duration = max(durations) if durations else 0.0
            
            # Countermovement je detektovan samo ako su zadovoljena SVA tri kriterijuma:
            # 1. Značajno spuštanje (brzina ILI displacement)
            # 2. Značajno smanjenje sile
            # 3. Trajanje > 50ms (ako je brzina kriterijum)
            significant_descent = (min_vel < cm_vel_threshold) or (min_disp < cm_disp_threshold)
            significant_force_drop = min_force < cm_force_threshold
            sufficient_duration = (min_vel >= cm_vel_threshold) or (cm_duration >= cm_duration_threshold)
            
            cm_detected = significant_descent and significant_force_drop and sufficient_duration
            
            if cm_detected:
                qc_flags['has_countermovement'] = True
                qc_flags['notes'] = f'Countermovement detected: v_min={min_vel:.3f}, d_min={min_disp:.3f}, F_min={min_force:.1f}'
        
    # G. Calculate Metrics
    m = {} 
    m['BW'] = bw
    m['BM'] = bm
    
    # VAŽNO: Za CMJ, redosled je: A (onset) → B (unweighting min) → C (propulsion peak) → D/F (takeoff)
    # Za SJ bez CM, B ne postoji, pa koristimo C direktno
    if jump_type == 2 and idx.get('B', -1) >= 0:  # CMJ sa unweighting min
        # Unweighting Phase: od onset-a do unweighting min-a
        m['dt_UP'] = (idx['B'] - idx['A']) / fs if idx['B'] > idx['A'] else 0
        # Braking Phase: od unweighting min-a do propulsion peak-a
        m['dt_BP'] = (idx['C'] - idx['B']) / fs if idx['C'] > idx['B'] else 0
        # Propulsion Phase: od propulsion peak-a do takeoff-a
        m['dt_PP'] = (idx['F'] - idx['C']) / fs if idx['F'] > idx['C'] else 0
        
        m['Fmin_UP'] = np.min(f_crop[idx['A']:idx['B']]) if idx['B'] > idx['A'] else 0
        m['Fmax_BP'] = np.max(f_crop[idx['B']:idx['C']]) if idx['C'] > idx['B'] else 0
        m['Fmax_PP'] = np.max(f_crop[idx['C']:idx['F']]) if idx['F'] > idx['C'] else 0
        # Srednje sile u fazama
        m['Favg_UP'] = np.mean(f_crop[idx['A']:idx['B']]) if idx['B'] > idx['A'] else 0
        m['Favg_BP'] = np.mean(f_crop[idx['B']:idx['C']]) if idx['C'] > idx['B'] else 0
        m['Favg_PP'] = np.mean(f_crop[idx['C']:idx['F']]) if idx['F'] > idx['C'] else 0
    else:  # SJ ili CMJ bez validnog B
        # Za SJ, koristimo C (propulsion peak) direktno
        m['dt_UP'] = (idx['C'] - idx['A']) / fs if idx['C'] > idx['A'] else 0
        m['dt_BP'] = 0  # Nema braking phase za SJ
        m['dt_PP'] = (idx['F'] - idx['C']) / fs if idx['F'] > idx['C'] else 0
        
        m['Fmin_UP'] = np.min(f_crop[idx['A']:idx['C']]) if idx['C'] > idx['A'] else 0
        m['Fmax_BP'] = 0  # Nema braking phase za SJ
        m['Fmax_PP'] = np.max(f_crop[idx['C']:idx['F']]) if idx['F'] > idx['C'] else 0
        # Srednje sile u fazama
        m['Favg_UP'] = np.mean(f_crop[idx['A']:idx['C']]) if idx['C'] > idx['A'] else 0
        m['Favg_BP'] = 0  # Nema braking phase za SJ
        m['Favg_PP'] = np.mean(f_crop[idx['C']:idx['F']]) if idx['F'] > idx['C'] else 0
    
    m['dt_FP'] = (idx['H'] - idx['F']) / fs if idx['H'] > idx['F'] else 0
    m['F_impact'] = np.max(f_crop[idx['H']:idx['H']+200]) if len(f_crop) > idx['H']+200 else 0
    
    # vTO mora biti brzina u trenutku odskoka (TO tački)
    if idx['F'] < len(vel):
        v_to_raw = vel[idx['F']]
    else:
        v_to_raw = 0.0
    
    # Validacija vTO
    # 1. Mora biti pozitivan
    # 2. Mora biti razuman (npr. < 10 m/s za SJ, < 15 m/s za CMJ)
    max_reasonable_vto = 10.0 if jump_type == 1 else 15.0
    
    if v_to_raw < 0:
        qc_flags['negative_vto'] = True
        if qc_flags['notes']:
            qc_flags['notes'] += '; Negative vTO'
        else:
            qc_flags['notes'] = 'Negative vTO'
        # Koristi vTO direktno (može biti negativan ako je greška u detekciji)
        m['v_to'] = v_to_raw
    elif v_to_raw > max_reasonable_vto:
        # Ekstremno visoka brzina - verovatno greška u drift correction ili event detection
        qc_flags['invalid_events'] = True
        if qc_flags['notes']:
            qc_flags['notes'] += f'; Extreme vTO ({v_to_raw:.2f} m/s)'
        else:
            qc_flags['notes'] = f'Extreme vTO ({v_to_raw:.2f} m/s)'
        # Koristi vTO direktno (označen kao neispravan)
        m['v_to'] = v_to_raw
    else:
        m['v_to'] = v_to_raw
    
    m['h_to_v'] = (m['v_to']**2) / (2 * CONFIG['GRAVITY']) 
    m['h_to_t'] = CONFIG['GRAVITY'] * (m['dt_FP']**2) / 8 
    m['h_max_disp'] = np.max(disp) if len(disp) > 0 else 0
    
    # Najniža tačka u Take-off fazi (displacement minimum u TO fazi)
    if idx['F'] > idx['A'] and idx['F'] <= len(disp):
        m['hmin_TO'] = np.min(disp[idx['A']:idx['F']])
    else:
        m['hmin_TO'] = 0
    
    # Impulse calculations - MORAJU biti pre h_to_impulse jer ga koristi
    if idx['F'] > idx['A'] and idx['F'] < len(f_crop) and idx['A'] >= 0:
        segment_tot = f_crop[idx['A']:idx['F']] - bw
        t_segment_tot = t_crop[idx['A']:idx['F']]
        if len(segment_tot) > 0 and len(t_segment_tot) > 0:
            result = cumulative_trapezoid(segment_tot, t_segment_tot)
            if len(result) > 0:
                m['J_tot'] = result[-1]
            else:
                m['J_tot'] = 0
        else:
            m['J_tot'] = 0
    else:
        m['J_tot'] = 0
    
    # Visina iz impulse-momentum teoreme: h = (J_tot / BM)^2 / (2*g)
    # gde je J_tot = integral od (F - BW) dt = integral od (F - BW) dt = BM * integral od acc dt
    # v_to = integral od acc dt (od onset-a do takeoff-a)
    # h_to_impulse = (J_tot / BM)^2 / (2*g) = (v_to_impulse)^2 / (2*g)
    # gde je v_to_impulse = J_tot / BM
    if m['J_tot'] != 0 and bm > 0:
        v_to_impulse = m['J_tot'] / bm
        m['h_to_impulse'] = (v_to_impulse**2) / (2 * CONFIG['GRAVITY'])
    else:
        m['h_to_impulse'] = 0
    
    # hmin_BP: minimalni displacement tokom braking phase (između unweighting min i propulsion peak)
    if jump_type == 2 and idx.get('B', -1) >= 0 and idx['C'] > idx['B']:
        # Za CMJ: braking phase je između B i C
        if idx['C'] <= len(disp):
            m['hmin_BP'] = np.min(disp[idx['B']:idx['C']])
        else:
            m['hmin_BP'] = 0
    else:
        # Za SJ: koristi ceo period od A do F
        if idx['F'] > idx['A'] and idx['F'] <= len(disp):
            m['hmin_BP'] = np.min(disp[idx['A']:idx['F']])
        else:
            m['hmin_BP'] = 0
    
    m['Pmax_PP'] = np.max(power[idx['C']:idx['F']]) if idx['F'] > idx['C'] else 0
    m['Pavg_PP'] = np.mean(power[idx['C']:idx['F']]) if idx['F'] > idx['C'] else 0
    m['RSI'] = m['h_to_v'] / (m['dt_PP'] + m['dt_BP'] + m['dt_UP']) if (m['dt_PP'] > 0) else 0
    
    # LS (Leg Stiffness): Fmax_BP / displacement change tokom braking phase
    if jump_type == 2 and idx.get('B', -1) >= 0 and idx['C'] > idx['B']:
        # Za CMJ: braking phase je između B i C
        d_disp_brake = abs(disp[idx['C']] - disp[idx['B']])
        m['LS'] = m['Fmax_BP'] / d_disp_brake if d_disp_brake > 0.005 else 0
    else:
        # Za SJ: koristi C i D (ali D = F za SJ)
        d_disp_brake = abs(disp[idx['F']] - disp[idx['C']]) if idx['F'] > idx['C'] else 0
        m['LS'] = m['Fmax_BP'] / d_disp_brake if d_disp_brake > 0.005 else 0
    
    # J_UP: Impulse tokom unweighting phase (od onset-a do unweighting min-a za CMJ, ili do propulsion peak-a za SJ)
    if jump_type == 2 and idx.get('B', -1) >= 0:
        # Za CMJ: unweighting phase je između A i B
        if idx['B'] > idx['A'] and idx['B'] < len(f_crop) and idx['A'] >= 0:
            segment_up = f_crop[idx['A']:idx['B']] - bw
            t_segment = t_crop[idx['A']:idx['B']]
            if len(segment_up) > 0 and len(t_segment) > 0:
                result = cumulative_trapezoid(segment_up, t_segment)
                if len(result) > 0:
                    m['J_UP'] = result[-1]
                else:
                    m['J_UP'] = 0
            else:
                m['J_UP'] = 0
        else:
            m['J_UP'] = 0
    else:
        # Za SJ: koristi A do C (propulsion peak)
        if idx['C'] > idx['A'] and idx['C'] < len(f_crop) and idx['A'] >= 0:
            segment_up = f_crop[idx['A']:idx['C']] - bw
            t_segment = t_crop[idx['A']:idx['C']]
            if len(segment_up) > 0 and len(t_segment) > 0:
                result = cumulative_trapezoid(segment_up, t_segment)
                if len(result) > 0:
                    m['J_UP'] = result[-1]
                else:
                    m['J_UP'] = 0
            else:
                m['J_UP'] = 0
        else:
            m['J_UP'] = 0
    
    # J_BP: Impulse tokom braking phase (između unweighting min i propulsion peak za CMJ)
    if jump_type == 2 and idx.get('B', -1) >= 0 and idx['C'] > idx['B']:
        # Za CMJ: braking phase je između B i C
        if idx['C'] < len(f_crop) and idx['B'] >= 0:
            segment_bp = f_crop[idx['B']:idx['C']] - bw
            t_segment = t_crop[idx['B']:idx['C']]
            if len(segment_bp) > 0 and len(t_segment) > 0:
                result = cumulative_trapezoid(segment_bp, t_segment)
                if len(result) > 0:
                    m['J_BP'] = result[-1]
                else:
                    m['J_BP'] = 0
            else:
                m['J_BP'] = 0
        else:
            m['J_BP'] = 0
    else:
        m['J_BP'] = 0  # Nema braking phase za SJ
    
    # J_PP: Impulse tokom propulsion phase (između propulsion peak i takeoff)
    if idx['F'] > idx['C'] and idx['F'] < len(f_crop) and idx['C'] >= 0:
        segment_pp = f_crop[idx['C']:idx['F']] - bw
        t_segment = t_crop[idx['C']:idx['F']]
        if len(segment_pp) > 0 and len(t_segment) > 0:
            result = cumulative_trapezoid(segment_pp, t_segment)
            if len(result) > 0:
                m['J_PP'] = result[-1]
            else:
                m['J_PP'] = 0
        else:
            m['J_PP'] = 0
    else:
        m['J_PP'] = 0
    
    # J_LAND: Impulse u landing fazi (od landing-a do najniže tačke - idx['min'])
    if idx.get('min', -1) >= 0 and idx['H'] >= 0 and idx['min'] > idx['H']:
        # Landing faza je od H (landing) do min (najniža tačka posle landing-a)
        if idx['min'] < len(f_crop) and idx['H'] >= 0:
            segment_land = f_crop[idx['H']:idx['min']+1] - bw
            t_segment = t_crop[idx['H']:idx['min']+1]
            if len(segment_land) > 0 and len(t_segment) > 0:
                result = cumulative_trapezoid(segment_land, t_segment)
                if len(result) > 0:
                    m['J_LAND'] = result[-1]
                else:
                    m['J_LAND'] = 0
            else:
                m['J_LAND'] = 0
        else:
            m['J_LAND'] = 0
    else:
        m['J_LAND'] = 0
    
    if idx['H'] >= 0 and idx['H'] < len(f_crop) and len(f_crop) > idx['H'] + 50:
        rfd_slice = np.diff(f_crop[idx['H']:idx['H']+50]) * fs
        if len(rfd_slice) > 0:
            m['RFD_Land'] = np.max(rfd_slice)
        else:
            m['RFD_Land'] = 0
    else:
        m['RFD_Land'] = 0
    
    # Dodaj QC flagove u metrike
    m['has_countermovement'] = qc_flags['has_countermovement']
    m['negative_vto'] = qc_flags['negative_vto']
    m['invalid_events'] = qc_flags['invalid_events']
    m['invalid_jump'] = qc_flags['has_countermovement'] or qc_flags['negative_vto'] or qc_flags['invalid_events'] or (m['v_to'] <= 0)
    m['qc_notes'] = qc_flags['notes']
    
    # Optional export internog signala za dodatne analize/plotovanje.
    # Podrazumevano je isključeno da bi postojeći tok ostao nepromenjen.
    if return_timeseries:
        m['_timeseries'] = {
            'time': t_crop,
            'force': f_crop,
            'force_left': fL_crop,
            'force_right': fR_crop,
            'force_full': force,
            'force_left_full': force_L,
            'force_right_full': force_R,
            'acc': acc,
            'vel': vel,
            'disp': disp,
            'idx': idx,
            'idx_abs': {
                'A': idx_A_abs,
                'F': idx_F_abs,
                'H': idx_H_abs,
                'min': idx_abs_min,
                'propulsion_peak': propulsion_peak_abs,
                'landing_peak': landing_peak_abs,
                'B': unweighting_min_abs if 'unweighting_min_abs' in locals() else None
            },
            'start_crop': start_crop,
            'end_crop': end_crop,
            'fs': fs
        }
    
    return m


def parse_filename(filename):
    """Parse filename format: ##_#_#.txt"""
    basename = Path(filename).stem
    pattern = r'^(\d+)_(\d+)_(\d+)$'
    match = re.match(pattern, basename)
    
    if not match:
        return None
    
    subject_id = match.group(1)
    jump_type_code = int(match.group(2))
    trial_no = int(match.group(3))
    
    # Mapiranje: 3 = SJ, 4 = CMJ
    if jump_type_code == 3:
        jump_type = 1  # SJ u VJ_1_1.py logici
    elif jump_type_code == 4:
        jump_type = 2  # CMJ u VJ_1_1.py logici
    else:
        return None
    
    return {
        'SubjectID': subject_id,
        'JumpTypeCode': jump_type_code,
        'TrialNo': trial_no,
        'JumpType': jump_type,
        'basename': basename,
        'filename': filename
    }


def process_force_plate_files(fp_dir: Path, jump_type_name: str):
    """Process all force plate files in a directory."""
    results = []
    
    files = sorted(fp_dir.glob("*.txt"))
    print(f"\n[{jump_type_name}] Pronađeno {len(files)} fajlova")
    
    for filepath in files:
        file_info = parse_filename(filepath.name)
        if file_info is None:
            print(f"  [SKIP] Nevalidan format imena: {filepath.name}")
            continue
        
        fL, fR = read_force_file(filepath)
        if fL is None or len(fL) == 0:
            print(f"  [SKIP] Ne može da se učita: {filepath.name}")
            continue
        
        # Calculate Total Force
        force_tot = fL + fR
        
        try:
            metrics = analyze_jump(force_tot, fL, fR, CONFIG['SAMPLE_RATE_AMTI'], 
                                   file_info['JumpType'], file_info['SubjectID'])
        except IndexError as e:
            import traceback
            tb_lines = traceback.format_exc().split('\n')
            error_line = [line for line in tb_lines if 'calculate_fp_kpis.py' in line and 'line' in line]
            print(f"  [ERROR] {filepath.name}: {e}")
            if error_line:
                print(f"  [LOCATION] {error_line[-1].strip()}")
            continue
        except Exception as e:
            import traceback
            print(f"  [ERROR] {filepath.name}: {e}")
            continue
        
        # Proveri da li je analiza uspešna (možda je BW invalid)
        if metrics is None:
            print(f"  [SKIP] {filepath.name}: Invalid BW or analysis failed")
            continue

        # SJ sa countermovement-om: iskljuci iz analize
        if file_info['JumpType'] == 1 and metrics.get('has_countermovement', False):
            print(f"  [SKIP] {filepath.name}: SJ sa countermovement-om - iskljuceno iz analize")
            continue
        
        # Kreiraj red za Excel
        row = {
            'FileName': filepath.name,
            'TrialID': file_info['basename'],
            'SubjectID': file_info['SubjectID'],
            'TrialNo': file_info['TrialNo'],
            'BW_N': metrics['BW'],
            'BM_kg': metrics['BM'],
            'dt_UP_s': metrics['dt_UP'],
            'Fmin_UP_N': metrics['Fmin_UP'],
            'Impulse_UP_Ns': metrics['J_UP'],
            'dt_BP_s': metrics['dt_BP'],
            'Depth_Max_m': metrics['hmin_BP'],
            'Fmax_BP_N': metrics['Fmax_BP'],
            'Stiffness_Nm': metrics['LS'],
            'dt_PP_s': metrics['dt_PP'],
            'V_Takeoff_ms': metrics['v_to'],
            'Height_V_m': metrics['h_to_v'],
            'Height_Impulse_m': metrics['h_to_impulse'],
            'Fmax_PP_N': metrics['Fmax_PP'],
            'Power_Max_W': metrics['Pmax_PP'],
            'Power_Avg_W': metrics['Pavg_PP'],
            'RSI': metrics['RSI'],
            'dt_Flight_s': metrics['dt_FP'],
            'Height_T_m': metrics['h_to_t'],
            'Impulse_Tot_Ns': metrics['J_tot'],
            'Impact_Force_N': metrics['F_impact'],
            'RFD_Landing_Ns': metrics['RFD_Land'],
            'has_countermovement': metrics.get('has_countermovement', False),
            'negative_vto': metrics.get('negative_vto', False),
            'invalid_bw': metrics.get('invalid_bw', False),
            'invalid_jump': metrics.get('invalid_jump', False),
            'qc_notes': metrics.get('qc_notes', ''),
        }
        
        results.append(row)
        print(f"  [OK] {filepath.name}")
    
    return results


def main():
    from paths_config import SJ_FORCE_PLATES, CMJ_FORCE_PLATES, EXCEL_DIR
    sj_fp_dir = SJ_FORCE_PLATES
    cmj_fp_dir = CMJ_FORCE_PLATES
    excel_file = EXCEL_DIR / "MoCap_KPIs.xlsx"
    
    print("=" * 90)
    print("IZRACUNAVANJE KPIs IZ FORCE PLATE PODATAKA")
    print("=" * 90)
    print(f"Vreme: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Excel fajl: {excel_file}")
    print("=" * 90)
    
    # Procesiraj SJ Force Plate fajlove
    sj_results = []
    if sj_fp_dir.exists():
        sj_results = process_force_plate_files(sj_fp_dir, "SJ_ForcePlates")
    else:
        print(f"\n[WARNING] Folder ne postoji: {sj_fp_dir}")
    
    # Procesiraj CMJ Force Plate fajlove
    cmj_results = []
    if cmj_fp_dir.exists():
        cmj_results = process_force_plate_files(cmj_fp_dir, "CMJ_ForcePlates")
    else:
        print(f"\n[WARNING] Folder ne postoji: {cmj_fp_dir}")
    
    # Učitaj postojeći Excel fajl
    if not excel_file.exists():
        print(f"\n[ERROR] Excel fajl ne postoji: {excel_file}")
        return 1
    
    # Učitaj sve postojeće sheetove
    excel_data = {}
    with pd.ExcelFile(excel_file) as xls:
        for sheet_name in xls.sheet_names:
            excel_data[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
    
    # Dodaj nove sheetove
    if sj_results:
        sj_df = pd.DataFrame(sj_results)
        # Sortiraj po SubjectID pa TrialNo
        sj_df = sj_df.sort_values(['SubjectID', 'TrialNo'], na_position='last')
        excel_data['SJ_FP'] = sj_df
    
    if cmj_results:
        cmj_df = pd.DataFrame(cmj_results)
        # Sortiraj po SubjectID pa TrialNo
        cmj_df = cmj_df.sort_values(['SubjectID', 'TrialNo'], na_position='last')
        excel_data['CMJ_FP'] = cmj_df
    
    # Sačuvaj ažurirani Excel fajl
    # Koristi privremeni fajl ako je originalni otvoren
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    import shutil
    
    temp_file = excel_file.parent / f"temp_{excel_file.name}"
    
    try:
        # Učitaj postojeći workbook
        wb = load_workbook(excel_file)
    except Exception as e:
        print(f"[ERROR] Ne može da se učita Excel fajl (možda je otvoren): {e}")
        print("Molimo zatvorite Excel fajl i pokrenite ponovo.")
        return 1
    
    # Dodaj nove sheetove
    for sheet_name, df in excel_data.items():
        if sheet_name in wb.sheetnames:
            # Ako sheet već postoji, obriši ga i kreiraj novi
            wb.remove(wb[sheet_name])
        
        # Kreiraj novi sheet
        ws = wb.create_sheet(sheet_name)
        
        # Dodaj header
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Dodaj podatke
        for r_idx, row in enumerate(df.itertuples(index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx+1, column=c_idx, value=value)
        
        # Freeze top row
        ws.freeze_panes = 'A2'
    
    # Sačuvaj u privremeni fajl prvo
    try:
        wb.save(temp_file)
        # Zameni originalni fajl
        shutil.move(str(temp_file), str(excel_file))
    except Exception as e:
        print(f"[ERROR] Ne može da se sačuva Excel fajl: {e}")
        if temp_file.exists():
            temp_file.unlink()
        return 1
    
    # Finalni izveštaj
    print("\n" + "=" * 90)
    print("FINALNI IZVESTAJ")
    print("=" * 90)
    print(f"Vreme završetka: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"\nSJ_FP: {len(sj_results)} fajlova obrađeno")
    print(f"CMJ_FP: {len(cmj_results)} fajlova obrađeno")
    print(f"\n[SUCCESS] Excel fajl ažuriran: {excel_file}")
    print(f"\nNovi sheetovi:")
    if sj_results:
        print(f"  - SJ_FP ({len(sj_results)} redova)")
    if cmj_results:
        print(f"  - CMJ_FP ({len(cmj_results)} redova)")
    
    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
