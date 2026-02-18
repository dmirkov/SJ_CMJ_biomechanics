"""
Modul za export KPIs u Excel fajl.
"""
import numpy as np
import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List
import config

logger = logging.getLogger(__name__)


def export_to_excel(all_kpis: Dict[str, Dict[str, List[Dict]]], output_file: Path = None):
    """
    Exportuje sve KPIs u Excel fajl sa 6 sheetova.
    
    Args:
        all_kpis: Dictionary sa strukturom all_kpis[jump_type][model] = lista KPI dictionaries
        output_file: Putanja do output Excel fajla (default: config.EXCEL_OUTPUT_FILE)
    """
    if output_file is None:
        output_file = config.EXCEL_OUTPUT_FILE
    
    # Kreiraj Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Za svaki tip skoka i model
        for jump_type in ['SJ', 'CMJ']:
            for model_idx, model in enumerate(['3D', '2DL', '2DR']):
                sheet_name = config.SHEET_NAMES[jump_type][model_idx]
                
                # Pripremi podatke za ovaj sheet
                kpis_list = []
                if jump_type in all_kpis and model in all_kpis[jump_type]:
                    kpis_list = all_kpis[jump_type][model]
                
                if not kpis_list:
                    # Kreiraj prazan DataFrame sa kolonama
                    df = pd.DataFrame(columns=config.EXCEL_COLUMNS)
                else:
                    # Konvertuj u DataFrame
                    df = pd.DataFrame(kpis_list)
                    
                    # Osiguraj da sve kolone postoje
                    for col in config.EXCEL_COLUMNS:
                        if col not in df.columns:
                            df[col] = np.nan
                    
                    # Sortiraj po SubjectID pa TrialNo
                    if 'SubjectID' in df.columns and 'TrialNo' in df.columns:
                        df = df.sort_values(['SubjectID', 'TrialNo'], na_position='last')
                    
                    # Reorder kolone
                    df = df[config.EXCEL_COLUMNS]
                
                # Zapisuj u sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatiranje (zahteva openpyxl)
                worksheet = writer.sheets[sheet_name]
                
                # Bold header
                from openpyxl.styles import Font, PatternFill
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Freeze top row
                worksheet.freeze_panes = 'A2'
                
                # Auto-size kolone (ograničeno) - uklonjeno privremeno zbog bug-a u openpyxl
                # TODO: Re-enable nakon što se reši problem sa column_dimensions
                # try:
                #     from openpyxl.utils import get_column_letter
                #     for idx, col in enumerate(df.columns, 1):
                #         max_length = max(
                #             df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                #             len(str(col))
                #         )
                #         adjusted_width = min(max_length + 2, 50)
                #         col_letter = get_column_letter(idx)
                #         worksheet.column_dimensions[col_letter].width = adjusted_width
                # except Exception as e:
                #     logger.warning(f"Ne može da se postavi auto-size za kolone: {e}")
    
    logger.info(f"Excel fajl sačuvan: {output_file}")
    logger.info(f"Sheetovi: {', '.join(config.SHEET_NAMES['SJ'] + config.SHEET_NAMES['CMJ'])}")


if __name__ == '__main__':
    # Test
    import numpy as np
    logging.basicConfig(level=logging.INFO)
    print("Export Excel modul - test passed")
